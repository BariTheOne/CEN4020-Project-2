import sqlite3
import openpyxl
from datetime import datetime
import re

#-------------------------
#   AUXILLARY FUNCTIONS
#-------------------------
def _splitTimeRange(input : str) -> tuple[str]:
    #check if the string is empty first
    if input == None or input.isspace():
        return (None, None)
    
    #otherwise, split the time range into 2 times
    strSplit = input.split(" - ")
    time1 = datetime.strptime(strSplit[0], "%I:%M %p")
    time2 = datetime.strptime(strSplit[1], "%I:%M %p")
    return (time1.isoformat(), time2.isoformat())

def _splitTAs(input : str) -> list[str]:
    #check if the string is empty first
    if input == None or input.isspace():
        return None
    
    #otherwise, split the TAs
    reStr = re.sub(r"\s*\([0-9]+\)", "", input)
    strSplit = reStr.split("\n")

    #save the TAs to a result list
    results = []
    for name in strSplit:
        if not name.isspace() and name != "TBD" and name != "" and "See " not in name:
            results.append(name)

    #only return the list if a valid name remaind
    if len(results) > 0:
        return results
    else:
        return None
    
def _splitTAEmails(input : str) -> list[str]:
    #check if the string is empty first
    if input == None or input.isspace():
        return None
    
    #otherwise, split the TAs
    strSplit = input.split("\n")

    #save the TAs to a result list
    results = []
    for email in strSplit:
        if not email.isspace() and email != "":
            results.append(email)

    #only return the list if a valid name remaind
    if len(results) > 0:
        return results
    else:
        return None

def _safeListConcatenate(l1, l2) -> list:
    if l1 == None and l2 == None:
        return None
    elif l1 == None:
        return l2
    elif l2 == None:
        return l1
    else:
        return l1 + l2
    
def _formatName(name : str) -> str:
    #clean the name up
    cleanName = re.sub(r"\s+\.", "", name)

    #check if the name is in firstname lastname order
    if "," in cleanName:
        #the name is in lastname, firstname format
        strSplit = cleanName.split(",")
        firstname = strSplit[1].strip()
        lastname = strSplit[0].strip()
        return firstname + " " + lastname
    else:
        #the name is already in the correct format
        return cleanName

def _cleanColumnName(input : str) -> str:
    #clean the name up
    cleanName = re.sub(r"[^A-Z ]", "", input.upper())
    return cleanName
    
#-------------------------------------
#   CLASS DEFINITION: EXCEL MANAGER
#-------------------------------------

class ExcelManager:
    def __init__(self, excel_filepath : str):
        self._excel = openpyxl.load_workbook(excel_filepath)
        self._sheet = self._excel.active
        self._map = {_cleanColumnName(cell.value): cell.column - 1 for cell in self._sheet[1]}
        print(self._map)

    def _safeEnrollmentCount(self, row) -> int:
        if "ENROLLMENT" not in self._map.keys():
            return 0
        else:
            num = row[self._map["ENROLLMENT"]].value
            if not isinstance(num, int):
                return 0
            else:
                return num
    
    def _safeCapacity(self, row) -> int:
        if "CAPACITY" not in self._map.keys():
            return 30
        else:
            num = row[self._map["CAPACITY"]].value
            if not isinstance(num, int):
                return 30
            else:
                return num

    def generateReportFromSheet(self, file_name : str = "Excel_Conflict_Report", filepath : str = "") -> bool:
        """
        Generates a TXT report of all conflicts in the Excel sheet.
        Returns False if conflicts are found. True otherwise.
        """

        #conflict flag
        has_conflict = True

        #open the report TXT file
        with open(file_name + ".txt", "w", encoding="utf-8") as report:

            #Write header
            report.write(f"---EXCEL SHEET REPORT---\n")

            #Column check
            #Does it have all the columns the system needs?
            report.write("\n1. COLUMNS\n")

            keys = self._map.keys()
            expected_columns = [
                "CAMPUS",
                "CRSE LEVL",
                "CRSE SECTION",
                "CRN",
                "SUBJ",
                "CRSE NUMB",
                "CRSE TITLE",
                "ENROLLMENT",
                "CAPACITY",
                "MEETING DAYS",
                "MEETING TIMES",
                "MEETING ROOM",
            ]

            for column in expected_columns:
                if column not in keys:
                    report.write(f""" - This spreadsheet is missing a column called "{column}".\n""")
                    has_conflict = False
        
            #7 Required Fields Check
            #Do the rows have the required fields?
            report.write("\n2. REQUIRED FIELDS\n")

            required_fields = [
                "CRN",
                "CAMPUS",
                "CRSE LEVL",
                "CRSE SECTION",
                "SUBJ",
                "CRSE NUMB",
                "CRSE TITLE"
            ]
            for row, row_index in zip(self._sheet.iter_rows(2, self._sheet.max_row), range(2, self._sheet.max_row)):
                for field in required_fields:
                    if row[self._map[field]].value == None:
                        report.write(f"""\tERROR - Row {row_index} has no {field}.\n""")
                        has_conflict = False

            #Duplicate CRN Check
            #Are there 2 classes with the same CRN?
            report.write("\n3. DUPLICATE CRNs\n")

            for row1, row_index1 in zip(self._sheet.iter_rows(2, self._sheet.max_row), range(2, self._sheet.max_row)):
                for row2, row_index2 in zip(self._sheet.iter_rows(row_index1 + 1, self._sheet.max_row), range(row_index1 + 1, self._sheet.max_row)):
                    if row1[self._map["CRN"]].value == row2[self._map["CRN"]].value:
                        report.write(f"""\tERROR - Rows {row_index1} and {row_index2} have the same CRN [{row1[self._map["CRN"]].value}].\n""")
                        has_conflict = False

            #Same Time and Place Check
            #Are there 2 classes that have the same time and place?
            report.write("\n4. SAME PLACE AND TIME CONFLICTS\n")

            for row1, row_index1 in zip(self._sheet.iter_rows(2, self._sheet.max_row), range(2, self._sheet.max_row)):
                for row2, row_index2 in zip(self._sheet.iter_rows(row_index1 + 1, self._sheet.max_row), range(row_index1 + 1, self._sheet.max_row)):
                    if row1[self._map["MEETING DAYS"]].value == row2[self._map["MEETING DAYS"]].value and\
                            row1[self._map["MEETING TIMES"]].value == row2[self._map["MEETING TIMES"]].value and\
                            row1[self._map["MEETING ROOM"]].value == row2[self._map["MEETING ROOM"]].value and\
                            row1[self._map["MEETING ROOM"]].value != "OFFT OFF" and\
                            row1[self._map["MEETING ROOM"]].value != "TBAT TBA":
                        report.write(f"""\tWARNING - Rows {row_index1} and {row_index2} [CRNs: {row1[self._map["CRN"]].value}, {row2[self._map["CRN"]].value}] are at same time [{row1[self._map["MEETING DAYS"]].value}, {row1[self._map["MEETING TIMES"]].value}] in the same room [{row1[self._map["MEETING ROOM"]].value}].\n""")
                        has_conflict = False

            #Proper Day Check
            #Do the rows have the required fields?
            report.write("\n5. MEETING DAY CHECK\n")

            for row, row_index in zip(self._sheet.iter_rows(2, self._sheet.max_row), range(2, self._sheet.max_row)):
                if row[self._map["MEETING DAYS"]].value not in ["MW", "TR", "F", None]:
                    report.write(f"""\tWARNING - Row {row_index} [CRN: {row[self._map["CRN"]].value}] is not meeting on MW, TR, or F [currently {row[self._map["MEETING DAYS"]].value}].\n""")
                    has_conflict = False

            #Instructor Check
            #Do the rows have an instructor
            report.write("\n6. INSTRUCTOR CHECK\n")

            for row, row_index in zip(self._sheet.iter_rows(2, self._sheet.max_row), range(2, self._sheet.max_row)):
                if row[self._map["INSTRUCTOR"]].value == None:
                    report.write(f"""\tWARNING - Row {row_index} [CRN: {row[self._map["CRN"]].value}] has no instructor.\n""")
                    has_conflict = False

        #return the bool
        return has_conflict

    #the following methods are really to made the code easier to read.
    def _createTables(self, connection : sqlite3.Connection, cursor : sqlite3.Cursor) -> bool:
        """
        Creates new tables if they do not exist. Returns False on error. Does not close the connection!

        :connection: Connection to the database
        :cursor: Cursor of the connection
        """
        try:
            #create a table for courses if it doesn't exist
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS "Courses" (
                    "CRN"	INTEGER NOT NULL UNIQUE,
                    "Campus"    TEXT NOT NULL,
                    "Level"	INTEGER NOT NULL,
                    "Section"	TEXT NOT NULL,
                    "Subject"	TEXT NOT NULL,
                    "Number"	TEXT NOT NULL,
                    "Title"	TEXT NOT NULL,
                    "Enrollment"    INTEGER,
                    "Capacity"  INTEGER,
                    "MeetingDays"	INTEGER,
                    "MeetingStartTime"	TEXT,
                    "MeetingEndTime"	TEXT,
                    "MeetingRoom"	TEXT,
                    PRIMARY KEY("CRN")
                );
                """
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False
        
        try:
            #create a table for instructors if it doesn't exist
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS "Instructors" (
                    "ID"	INTEGER NOT NULL UNIQUE,
                    "Name"	TEXT NOT NULL,
                    "Email"	TEXT    UNIQUE,
                    PRIMARY KEY("ID" AUTOINCREMENT)
                );
                """
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False

        try:
            #create a junction table for instructor-course pairs if it doesn't exist
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS "CourseInstructor" (
                    "ID"    INTEGER NOT NULL UNIQUE,
                    "CRN"	INTEGER NOT NULL,
                    "InstructorID"	INTEGER NOT NULL,
                    CONSTRAINT "CRN" FOREIGN KEY("CRN") REFERENCES "Courses"("CRN"),
                    CONSTRAINT "InstructorID" FOREIGN KEY("InstructorID") REFERENCES "Instructors"("ID")
                    PRIMARY KEY("ID" AUTOINCREMENT)
                );
                """
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False

        try:
            #create a table for TAs if it doesn't exist
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS "TeachingAssistants" (
                    "ID"	INTEGER NOT NULL UNIQUE,
                    "Name"	TEXT NOT NULL,
                    "Email" TEXT    UNIQUE,
                    "Level" TEXT,
                    PRIMARY KEY("ID" AUTOINCREMENT)
                );
                """
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False

        try:
            #create a junction table for TA-course pairs if it doesn't exist
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS "CourseTA" (
                    "ID"    INTEGER NOT NULL UNIQUE,
                    "CRN"	INTEGER NOT NULL,
                    "TAID"	INTEGER NOT NULL,
                    CONSTRAINT "CRN" FOREIGN KEY("CRN") REFERENCES "Courses"("CRN"),
                    CONSTRAINT "TAID" FOREIGN KEY("TAID") REFERENCES "TeachingAssistants"("ID")
                    PRIMARY KEY("ID" AUTOINCREMENT)
                );
                """
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False
        
        try:
            #create a table for studentss if it doesn't exist
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS "Students" (
                    "ID"	INTEGER NOT NULL UNIQUE,
                    "Name"	TEXT NOT NULL,
                    "Email" TEXT    UNIQUE,
                    PRIMARY KEY("ID" AUTOINCREMENT)
                );
                """
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False

        try:
            #create a junction table for student-course pairs if it doesn't exist
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS "CourseStudent" (
                    "ID"    INTEGER NOT NULL UNIQUE,
                    "CRN"	INTEGER NOT NULL,
                    "StudentID"	INTEGER NOT NULL,
                    CONSTRAINT "CRN" FOREIGN KEY("CRN") REFERENCES "Courses"("CRN"),
                    CONSTRAINT "StudentID" FOREIGN KEY("StudentID") REFERENCES "Student"("UNumber")
                    PRIMARY KEY("ID" AUTOINCREMENT)
                );
                """
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False
        
        #commit the changes
        connection.commit()
        return True
    
    def _insertCourse(self, connection : sqlite3.Connection, cursor : sqlite3.Cursor, course_data : tuple) -> bool:        
        #add the data through a prepared statement
        try: 
            cursor.execute(
                """
                INSERT INTO "Courses" VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT("CRN") DO NOTHING;
                """,
                course_data
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False
        
        #commit the changes
        connection.commit()
        return True
            
    def _insertInstructor(self, connection : sqlite3.Connection, cursor : sqlite3.Cursor, instructor_data : tuple) -> bool:
        #add the data through a prepared statement
        try: 
            cursor.execute(
                """
                INSERT INTO "Instructors" (Name, Email) VALUES (?, ?)
                ON CONFLICT DO NOTHING;
                """,
                instructor_data
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False
        
        #commit the changes
        connection.commit()
        return True

    def _insertTA(self, connection : sqlite3.Connection, cursor : sqlite3.Cursor, ta_data : tuple) -> bool:
        #add the data through a prepared statement
        try: 
            cursor.execute(
                """
                INSERT INTO "TeachingAssistants" (Name, Email, Level) VALUES (?, ?, ?)
                ON CONFLICT DO NOTHING;
                """,
                ta_data
            )
        except sqlite3.Error as error:
            print("Error!:", error)
            return False
        
        #commit the changes
        connection.commit()
        return True

    #ADD TO DATABASE FROM SHEET
    def addToDatabaseFromSheet(self, database_name : str, database_filepath = ""):
        
        #check if the spreadsheet is valid
        if self.generateReportFromSheet() == False:
            print("Error!: This spreadsheet has issues. Please look at the report for details.\nAborting...")
            return
        
        #get the full path
        full_path = database_filepath + database_name + ".db"

        #establish an SQL connection
        connection = sqlite3.connect(full_path)
        cursor = connection.cursor()

        #create lists for info we have already seen to avoid duplicates
        seen_instructors = []
        seen_tas = []

        #create tables
        self._createTables(connection, cursor)

        #add course details
        for row in self._sheet.iter_rows(2, self._sheet.max_row):
            #organize the course data
            course_data = (
                row[self._map["CRN"]].value,   #CRN
                row[self._map["CAMPUS"]].value,   #Campus
                row[self._map["CRSE LEVL"]].value,   #Level
                row[self._map["CRSE SECTION"]].value,   #Section
                row[self._map["SUBJ"]].value,   #Subject
                row[self._map["CRSE NUMB"]].value,   #Number
                row[self._map["CRSE TITLE"]].value,   #Title
                self._safeEnrollmentCount(row),   #Enrollment
                self._safeCapacity(row),            #Capacity
                row[self._map["MEETING DAYS"]].value,  #MeetingDays
                _splitTimeRange(row[self._map["MEETING TIMES"]].value)[0],               #MeetingStartTime
                _splitTimeRange(row[self._map["MEETING TIMES"]].value)[1],               #MeetingEndTime
                row[self._map["MEETING ROOM"]].value,  #MeetingRoom
            )

            #add the data
            if self._insertCourse(connection, cursor, course_data) == False:
                continue

            #check if there is an instructor for the class
            if row[self._map["INSTRUCTOR"]].value != None:
                #check if it's a new instructor
                if row[self._map["INSTRUCTOR"]].value not in seen_instructors:
                    #mark the instructor as seen
                    seen_instructors.append(row[self._map["INSTRUCTOR"]].value)

                    #organize the instructor data
                    instructor_data = (
                        _formatName(row[self._map["INSTRUCTOR"]].value),  #Name
                        row[self._map["INSTRUCTOR EMAIL"]].value,  #Email
                    )

                    #add the data
                    self._insertInstructor(connection, cursor, instructor_data)

                #insert a new entry to the junction table
                #get the ID of the instructor
                sql_results = cursor.execute(
                    """
                    SELECT "ID" FROM Instructors WHERE "Email" = ?;
                    """,
                    (row[self._map["INSTRUCTOR EMAIL"]].value,)
                )
                instructor_id = sql_results.fetchone()[0]

                #check if an course-instructor pair exists
                sql_results = cursor.execute(
                    """
                    SELECT "ID" FROM CourseInstructor WHERE (CRN, InstructorID) = (?, ?);
                    """,
                    (row[self._map["CRN"]].value, instructor_id)
                )
                course_instructor_pair = sql_results.fetchall()

                #if the pair does not exist, add it.
                if len(course_instructor_pair) == 0:

                    #organize the pair data
                    course_instructor_data =(
                        row[self._map["CRN"]].value,
                        instructor_id,
                    )

                    #add the data through a prepared statement
                    cursor.execute(
                        """
                        INSERT INTO "CourseInstructor" (CRN, InstructorID) VALUES (?, ?)
                        ON CONFLICT DO NOTHING;
                        """,
                        course_instructor_data
                    )

                    #commit the change
                    connection.commit()
            
            #check if there are TAs for the class
            ta_list = _safeListConcatenate(_splitTAs(row[self._map["UGTAS"]].value), _splitTAs(row[self._map["GRAD TAS"]].value))
            ta_email_list = _safeListConcatenate(_splitTAEmails(row[self._map["UGTA EMAILS"]].value), _splitTAEmails(row[self._map["GRAD TA EMAILS"]].value))
            if ta_list != None and ta_email_list != None:
                for ta_name, ta_email in zip(ta_list, ta_email_list):
                    if ta_name not in seen_tas:
                        #mark the instructor as seen
                        seen_tas.append(ta_name)

                        #see if the TA is UG or GR
                        ug_flag = False
                        if row[self._map["UGTAS"]].value != None and ta_name in row[self._map["UGTAS"]].value:
                            ug_flag = True

                        #organize the instructor data
                        ta_data = (
                            ta_name,
                            ta_email,
                            "UG" if ug_flag else "GR"
                        )

                        #add the data
                        self._insertTA(connection, cursor, ta_data)

                    #insert a new entry to the junction table
                    #get the ID of the TA
                    sql_results = cursor.execute(
                        """
                        SELECT "ID" FROM "TeachingAssistants" WHERE "Email" = ?;
                        """,
                        (ta_email,)
                    )
                    ta_id = sql_results.fetchone()[0]

                    #check if an course-TA pair exists
                    sql_results = cursor.execute(
                        """
                        SELECT "ID" FROM CourseTA WHERE (CRN, TAID) = (?, ?);
                        """,
                        (row[self._map["CRN"]].value, ta_id)
                    )
                    course_ta_pair = sql_results.fetchall()

                    #if the pair does not exist, add it.
                    if len(course_ta_pair) == 0:

                        #organize the pair data
                        course_ta_data =(
                            row[self._map["CRN"]].value,
                            ta_id,
                        )

                        #add the data through a prepared statement
                        cursor.execute(
                            """
                            INSERT INTO "CourseTA" (CRN, TAID) VALUES (?, ?)
                            ON CONFLICT DO NOTHING;
                            """,
                            course_ta_data
                        )

                        #commit the change
                        connection.commit()
        
        #close the connection
        connection.close()

#DEBUGGING
if __name__ == "__main__":
    em = ExcelManager("Bellini Classes S26.xlsx")
    em.addToDatabaseFromSheet("TEST")


    """
    df = openpyxl.load_workbook("Bellini Classes S25.xlsx")
    df1 = df.active

    for row in range(0, 4):
        for col in df1.iter_cols(1, df1.max_column):
            print(col[row].value)
    """